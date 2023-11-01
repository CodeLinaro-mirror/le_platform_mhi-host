import sys
import os
import re
import csv
import numpy as np
from itertools import islice
from matplotlib.colors import BoundaryNorm
from matplotlib.ticker import MaxNLocator
import matplotlib.pyplot as plt
import matplotlib.colors as colors
import logging
import math
import argparse
import time
from array import *
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


eye_width = 0
leftmargin = 0
rightmargin = 0
VOLTAGE_STEP_SIZE = 1.5 # 1.5mv step size
TIMING_STEP_SIZE = 0.03125 #0.03125 # 1/320.015625
delete_filename  = ""
lane_info = ""
exact_file = ""
left_EW = ""
right_EW = ""
Total_EW = ""
eye_height = ""


RESULTDIR = os.getcwd()
#RESULTDIR = RESULTDIR #+ "\\"

##############################################################################################
# Gen 4 Eye Mask specifications
# VRX-CH-EH-16G Eye height 15 (min) mVPP Eye height at BER=10-12. Note 1
# TRX-CH-EW-16G Eye width at zero crossing 0.3 (min) UI Eye width at BER=10-12
##############################################################################################
MIN_LEFT_MARGIN  = 0.15
MIN_RIGHT_MARGIN = 0.15
MIN_EYE_WIDTH    = 0.3
LEFT_MARGIN_BIT_POS = 2
RIGHT_MARGIN_BIT_POS = 1
EYEWIDTH_BIT_POS = 0

MIN_EYE_HEIGHT    = 15
EYEHEIGHT_BIT_POS = 0

MARGIN_RESULT_BIT_MASK = 0x1

Ny=128

'''
Error Count thresholds for both Timing and Voltage Margins
'''
TIMING_MARGIN_ERR_THRESHOLD = 5
VOLTAGE_MARGIN_ERR_THRESHOLD = 5

'''
Measure the Eye Height
'''
def measure_eye_height(data, mean_UI_center):
		vmargin_result = []

		'''
			 bit2 - Left margin result
			 bit1 - Right margin result
			 bit0 - Eye width result
		'''
		margin_result = MARGIN_RESULT_BIT_MASK
		ylen, xlen = data.shape

		'''
		Logic to find the Eye Height
		'''
		# Vertical margin
		#xcoord = int(xlen / 2) # vijraj commented this because this gives the wrong EH as we are not follwing max of Y(mean_UI_center-3, mean_UI_center+3)
		#below logic added by vijraj to satify the base spec for finding the EH
		eheight = array("i", [0,0,0,0,0,0,0])
		mean_UI_center = mean_UI_center - 3 # max of Y(mean_UI_center-3, mean_UI_center+3)
		for i in range(7):

				ycoord = 0
				while ycoord < ylen:
						#if data[ycoord, xcoord] <= VOLTAGE_MARGIN_ERR_THRESHOLD: # this is exactly used to capture at x=32 and this is not as per base spec
						if data[ycoord, mean_UI_center] > VOLTAGE_MARGIN_ERR_THRESHOLD: # the EH is calculating at mean of the EW, this needs to be reviewd with kislay/shivam
									break
						eheight[i] = ycoord
						ycoord += 1
				mean_UI_center = mean_UI_center + 1 # max of Y(avg-3, avg+3)

		maximum = max(eheight)

		meye_height = maximum * VOLTAGE_STEP_SIZE


		'''
		Check if the meye_height is meeting the specification
		'''
		if meye_height > MIN_EYE_HEIGHT:
				margin_result = margin_result & ~(1 << EYEHEIGHT_BIT_POS)



		'''
		Declare Pass/Fail for the Eye Height
		Check Left, Right margins and eye width if they meet the Gen 4 specification
		Section 8.5.1.6 Pass/Fail Eye Characteristics PCIe Base Gen 4 specification
		'''
		if margin_result == 0:
				mresult = 'PASS'
		else:
				mresult = 'FAIL'

		vmargin_result.append(meye_height)
		vmargin_result.append(mresult)

		return vmargin_result

'''
Measure the Eye Width
'''
def measure_eye_width(data):
		result = []
		'''
			 bit2 - Left margin result
			 bit1 - Right margin result
			 bit0 - Eye width result
		'''
		margin_result = MARGIN_RESULT_BIT_MASK
		ylen, xlen = data.shape
		ycoord = 0

		'''
		Logic to find the Eye Width and also the Left margin and the right margin
		'''
		#First find Left margin
		xcoord = int(xlen/2)
		lefty = 0
		while xcoord >= 0:
				if data[ycoord, xcoord] > TIMING_MARGIN_ERR_THRESHOLD:
						break
				lefty = xcoord
				xcoord -= 1

		leftmargin = lefty
		mleftmargin = int(xlen/2) - leftmargin
		mleftmargin = mleftmargin * TIMING_STEP_SIZE


		'''
		Check if the mleftmargin is meeting the specification
		'''
		if mleftmargin > MIN_LEFT_MARGIN:
				margin_result = margin_result & ~(1 << LEFT_MARGIN_BIT_POS)


		xcoord = int(xlen/2)

		righty = 0
		while xcoord < xlen:
				if data[ycoord, xcoord] > 5: #vijraj added
						break
				righty = xcoord
				xcoord += 1

		rightmargin = righty


		mrightmargin = rightmargin - int(xlen/2) + 1

		mrightmargin = mrightmargin * TIMING_STEP_SIZE


		'''
		Check if the mrightmargin is meeting the specification
		'''
		if mrightmargin > MIN_RIGHT_MARGIN:
				margin_result = margin_result & ~(1 << RIGHT_MARGIN_BIT_POS)


		meye_width = mrightmargin + mleftmargin

		sum = 0
		temp = 0
		for i in range(leftmargin,rightmargin):
				sum = sum + i
				temp = temp + 1
		if temp == 0:
				avg = 0
		else:
				avg = (sum/temp)

		ui_cent = (leftmargin+rightmargin)/2


		'''
		Logic to find the Eye Width and also the Left margin and the right margin
		'''
		#First find Left margin
		xcoord = int(avg) #int(xlen/2)
		lefty = 0
		while xcoord >= 0:
				if data[ycoord, xcoord] > 5:
						break
				lefty = xcoord
				xcoord -= 1

		leftmargin = lefty
		mleftmargin = int(avg) - leftmargin
		mleftmargin = mleftmargin * TIMING_STEP_SIZE


		'''
		Check if the mleftmargin is meeting the specification
		'''
		if mleftmargin > MIN_LEFT_MARGIN:
				margin_result = margin_result & ~(1 << LEFT_MARGIN_BIT_POS)


		xcoord = int(avg) #int(xlen/2)

		righty = 0
		while xcoord < xlen:
				if data[ycoord, xcoord] > 5: #vijraj added
						break
				righty = xcoord
				xcoord += 1

		rightmargin = righty



		mrightmargin = rightmargin - int(avg) + 1

		mrightmargin = mrightmargin * TIMING_STEP_SIZE


		'''
		Check if the mrightmargin is meeting the specification
		'''
		if mrightmargin > MIN_RIGHT_MARGIN:
				margin_result = margin_result & ~(1 << RIGHT_MARGIN_BIT_POS)


		meye_width = mrightmargin + mleftmargin
		'''
		Check if the meye_width is meeting the specification
		'''
		if meye_width > MIN_EYE_WIDTH:
				margin_result = margin_result & ~(1 << EYEWIDTH_BIT_POS)

		'''
		Declare Pass/Fail for the Eye width
		Check Left, Right margins and eye width if they meet the Gen 4 specification
		Section 8.5.1.6 Pass/Fail Eye Characteristics PCIe Base Gen 4 specification
		'''
		if margin_result == 0:
				mresult = 'PASS'
		else:
				mresult = 'FAIL'

		result.append(mleftmargin) # bit2 - Left margin result
		result.append(mrightmargin) # bit1 - Right margin result
		result.append(meye_width) # bit0 - Eye width result
		result.append(mresult)
		result.append(avg)

		return result


'''
		API to Extract set of N lines from the FILE
'''
def next_n_lines(file_ptr, N):
		return [x.strip() for x in islice(file_ptr, N)]


def parse_margining_log(inputfilepath, sequence, args):

		'''
		Headers for further parsing of the logs
		'''
		global delete_filename
		global lane_info
		global exact_file
		global left_EW
		global right_EW
		global Total_EW
		global eye_height
		header = ['L', 'x', 'y', 'E']
		configheader = ['Mode', 'LaneWidth']
		strx1 = 'pcie negotiated link speed'
		strx2 = 'pcie negotiated link width'
		strx4 = 'Margining'

		'''
		No. of lines to extract from log for each lane
		64 Timing Steps x 128 Vertical Steps = 8192
		'''
		num_lines_to_extract = 64*(Ny+4)       # For extracting N lines from log
		margin_files_list = []

		'''
		First extract the margining results for EP
		from the main log file and log them into individual log files
		for both RC and EP with lane combinations
		'''

		logpath = os.path.dirname(inputfilepath)
		with open(inputfilepath) as f:

				for line in f:

						if strx1 in line:
								logging.debug("Matched : %s ", strx1)
								genstr=re.split('[:\n]', line)
								continue

						if strx4 in line:
								logging.debug("Matched : %s", strx4)
								tokens = line.split()
								numtokens = len(line.split())
								lane_info = tokens[5]
								marginfile = tokens[2] + tokens[3] + tokens[4] + tokens[5]

								marginfile = marginfile.replace("," , "_") + ".txt"
								exact_file = marginfile
								delete_filename = marginfile
								margindata_lines = next_n_lines(f, num_lines_to_extract)
								filestr = logpath + "/" + marginfile
								margin_files_list.append(filestr)
								marginfileptr = open(filestr, 'w+')
								for marginlines in margindata_lines:
										marginfileptr.write(marginlines + "\n")
								marginfileptr.close()
								continue

		'''
		Now using these individual log files containing margining data for RC and EP
		with lane combinations extract the 'Lane', 'xcoord', 'ycoord', 'Errors'
		into a dictionary and save them as well
		'''
		for filestr in margin_files_list:
				with open(filestr) as f:
						with open(filestr + ".csv", "wt") as fw:
								csv_writer = csv.DictWriter(
										fw,
										fieldnames=header,
										restval="NA",
										extrasaction='ignore',  # ignore extra values in the dictionary
										quoting=csv.QUOTE_NONNUMERIC
								)

								csv_writer.writeheader()  # write header
								for line in f:
										if not line.strip():
												continue    #skip blank lines
										str1 = re.split(',|:|\n',line)

										row = str1[1] + ',' + str1[3] + ',' + str1[5] + ',' + str1[7] + '\n'
										fw.write(row)

		'''
		Now form the one dimensional array with margining results for each of the timing step
		in both positive and negative direction and then append to 2D array
		'''

		for filestr in margin_files_list:

				with open(filestr + ".csv") as fr:
						mm = np.zeros((Ny+4, 64)) #64np.zeros((28, 64))
						tmparray = []
						csv_reader = csv.DictReader(fr)
						for row in csv_reader:
								xcord=0
								tmparray.append(int(row['E'], 16))     # converting hex string to integer
								if int(row['x']) != 63:  #63
										xcord = xcord + 1
										continue
								xcord=0
								mm[int(row['y']),:] = tmparray[:]
								tmparray = []

						np.savetxt(filestr + ".log2", mm, delimiter=",", fmt='%d')

		'''
		Logic for extrapolating data in the files for plotting the Eye
		'''

		import time
		timestr = time.strftime("_%Y%m%d-%H%M%S")


		for filestr in margin_files_list:

				NUMX = 0.5
				NUMY = (Ny*3)/2 + 4  #36
				x = np.arange(-NUMX, NUMX, TIMING_STEP_SIZE)

				'''
				Changed as we are having single ended sweeping
				'''

				y = np.arange(0, NUMY, VOLTAGE_STEP_SIZE)

				xx, yy = np.meshgrid(x, y, sparse=True)
				data = np.genfromtxt(filestr + ".log2", delimiter=',')

				value_3=Ny #24
				value_4=0
				for value_3 in range(Ny,Ny+4):
						for value_4 in range(0,64):
								data[value_3][value_4]=60000

				data_new=data


				Data_new=[None]*64*Ny
				i=0
				j=0
				Ncount=0
				for i in range(0,Ny):
						for j in range(0,64):
								Data_new[Ncount]=data[i][j]
								Ncount=Ncount+1


				value_3=0
				value_4=0
				value_4_past=0
				pos_final=0
				divide_val=0

				for value_3 in range(0,Ny):
						for value_4 in range(0,64):
								if(data[value_3][value_4]<=5):
										if(value_4<32):
												value_4=value_4+64
										pos_final=pos_final+value_4
										divide_val=divide_val + 1


				pos_final=(pos_final/divide_val)
				pos_final_1=math.floor(pos_final)



				value_3=0
				value_4=0
				a=int(0)
				value_5=pos_final_1 - 32
				Ncount=pos_final_1 - 32
				Ncount_temp=0
				Ncount_orig=pos_final_1 - 32
				for value_3 in range(0,Ny):
						for value_4 in range(0,64):
								if((Ncount%64)==0):
										Ncount_orig=Ncount
										Ncount=Ncount_temp
										Ncount_temp=Ncount_temp + 64
								if(Ncount==Ncount_temp+25):
										Ncount=Ncount_orig

								a = int(Data_new[Ncount])
								data[value_3][value_4]=a
								Ncount=Ncount+1

				res = measure_eye_width(data)
				mean_UI = res[4]
				mean_UI = int(mean_UI)
				timing_result = res[3]
				temp = res[2] - res[1]
				if res[0] < res[1]:
						temp = -temp
				resultstr = "Timing Margins : Left = " + str(res[0]) + " UI, Right = " + str(res[1]) + " UI, Eye centre = " + str(temp) + " UI, Eye Width = "\
										+ str(res[2]) + " UI"
				left_EW = str(res[0])
				right_EW = str(res[1])
				Total_EW = str(res[2])
				res = measure_eye_height(data,mean_UI)
				resultstr = resultstr + "\n" + "Voltage Margins : Eye Height = " + str(res[0]) + " mV"
				eye_height = str(res[0])
				voltage_result = res[1]

				textclr = 'gray'
				if timing_result == 'PASS' and voltage_result == 'PASS':
						textclr = 'green'
						margin_result = 'PASS'
				if timing_result == 'FAIL' or voltage_result == 'FAIL':
						textclr = 'red'
						margin_result = 'FAIL'


				genstr_1="Margin"
				plottitle = filestr.replace(".txt", timestr)
				exact_file = exact_file.replace(".txt", timestr)
				imagename = plottitle + '_' + margin_result + ".png"

				plt.figure(imagename, figsize=(12,8))
				plottitle = plottitle.replace(RESULTDIR,"")
				plottitle = plottitle.replace("_log.csv", " ")
				plottitle = resultstr + " " + plottitle +" " +margin_result
				print("====== %s ======" % plottitle)


				value_3=0
				value_4_1=17
				value_4=0

				data_new_1=  np.zeros((Ny+4-1, 32))#[None]*[None]*24*32
				for value_3 in range(0,Ny):
						for value_4 in range(0,32):
							if(value_4_1==49):
								value_4_1=17

							data_new_1[value_3][value_4]=data_new[value_3][value_4_1]

							value_4_1=value_4_1 + 1

				value_3=Ny
				value_4=0
				for value_3 in range(Ny,Ny+3):
						for value_4 in range(0,32):
								data_new_1[value_3][value_4]=60000
				if(sequence == "p"):
					h = plt.pcolormesh(xx, yy, data_new_1,norm=colors.PowerNorm(gamma=0.10),
											 cmap="inferno_r") #cmap='PuBu_r  inferno_r'
					cbar = plt.colorbar(h)
				elif(sequence == "n"):
					h = plt.pcolormesh(xx, -yy, data_new_1,norm=colors.PowerNorm(gamma=0.10),
											 cmap="inferno_r") #cmap='PuBu_r  inferno_r'
					cbar = plt.colorbar(h)
					cbar.ax.invert_yaxis() #added by vijraj to invert the error bar

				cbar.set_label('No. of Errors', rotation=270, labelpad=20)

				####

				plt.minorticks_on()
				plt.grid(True, which='both', linestyle='-')
				plt.xlim(-0.5, 0.43)   ### For showing intervals in steps of 0.1 UI
				plt.title(plottitle, fontsize=14, color=textclr)
				plt.xlabel('Unit Intervals (UI)')
				plt.ylabel('milli volts (mV)')
				plt.text(-0.35, 3*Ny+Ny/3+3, resultstr, style='italic', weight='bold', color='white',
								bbox={'facecolor': 'red','alpha': 0.4, 'pad': 8})

				'''
				Plot the eye
				'''
				if(sequence == "p"):
						save_location = os.path.join(os.getcwd(),args.path,'positive_sequence',exact_file)
				elif(sequence == "n"):
						save_location = os.path.join(os.getcwd(),args.path,'negative_sequence', exact_file)
				print(save_location)
				plt.savefig(save_location)
				plt.close('all')



if __name__  ==  '__main__':

		'''
		if len(sys.argv) < 2:
				print("=====================================================================")
				print('Insufficient arguments, please specify filename with complete path')
				print("Example : Logparser_pciemargining.py C:\\TEST\\newtestlog.txt")
				print("=====================================================================")
				exit(0)
		'''

		parser = argparse.ArgumentParser(
										 description='Eye diagram Generator')
		parser.add_argument("-folder",
													dest="path",
													help="(REQUIRED) Sequence (p or n)",
													default="",
													required=True)
		args = parser.parse_args()
		timestr = time.strftime("%Y%m%d-%H%M%S")
		print(args.path)
		Excel_file_positive = f"{timestr}_positive.xlsx"
		Excel_file_negative = f"{timestr}_negative.xlsx"
		wb_pos = Workbook()
		wb_neg = Workbook()
		runs = 1
		for k in range (0,runs):

			for i in range(0,16):
				filename = os.path.join(RESULTDIR,args.path,'positive_sequence_logs',f"lane{i}.txt")
				with open(filename , 'r') as f:
					data = f.readlines()
					str0 = data[0]
					start = str0.find('*')
					data[0] = str0[start:]


				with open(filename , 'w') as f:
					f.writelines(data)

				parse_margining_log(filename, 'p',args)
				eye_params = [left_EW, right_EW , Total_EW, eye_height]

				ws = wb_pos.active

				ws.merge_cells(f'A{i*runs+2}:A{(i+1)*runs+1}')
				ws.cell(row = i*runs+2 , column = 1 , value = f"{i}")

				ws.cell(row = 1 , column = 1 , value = "Lane")
				ws.cell(row = 1 , column = 2 , value = "Run")
				ws.cell(row = 1 , column = 3 , value = "Left EW")
				ws.cell(row = 1 , column = 4 , value = "Right EW")
				ws.cell(row = 1 , column = 5 , value = "Total EW")
				ws.cell(row = 1 , column = 6 , value = "Eye H.")

				row_to_write = i*runs+2+k
				ws.cell(row = row_to_write, column = 2 , value = f"Run{k}")
				color = "ffffff" if (i%2) else "00ABF0"
				for j in range(3,7):
					cell = ws.cell(row = row_to_write , column = j)
					cell.value = eye_params[j-3]
					cell.font = Font(size = 11, name = 'Calibri')
					cell.fill = PatternFill("solid", start_color=color)

				print(f"+ve sequence image Generated for lane{lane_info}")

				try:
					os.remove(f"{delete_filename}")
					os.remove(f"{delete_filename}.csv")
					os.remove(f"{delete_filename}.log2")
				except:
					print("Some exceptions occurred")

			for i in range(0,16):
				filename = os.path.join(RESULTDIR,args.path,'negative_sequence_logs',f"lane{i}.txt")
				with open(filename , 'r') as f:
					data = f.readlines()
					str0 = data[0]
					start = str0.find('*')
					data[0] = str0[start:]


				with open(filename , 'w') as f:
					f.writelines(data)

				parse_margining_log(filename, 'n',args)
				eye_params = [left_EW, right_EW , Total_EW, eye_height]

				ws = wb_neg.active

				ws.merge_cells(f'A{i*runs+2}:A{(i+1)*runs+1}')
				ws.cell(row = i*runs+2 , column = 1 , value = f"{i}")

				ws.cell(row = 1 , column = 1 , value = "Lane")
				ws.cell(row = 1 , column = 2 , value = "Run")
				ws.cell(row = 1 , column = 3 , value = "Left EW")
				ws.cell(row = 1 , column = 4 , value = "Right EW")
				ws.cell(row = 1 , column = 5 , value = "Total EW")
				ws.cell(row = 1 , column = 6 , value = "Eye H.")

				row_to_write = i*runs+2+k
				ws.cell(row = row_to_write, column = 2 , value = f"Run{k}")
				color = "ffffff" if (i%2) else "00ABF0"
				for j in range(3,7):
					cell = ws.cell(row = row_to_write , column = j)
					cell.value = eye_params[j-3]
					cell.font = Font(size = 11, name = 'Calibri')
					cell.fill = PatternFill("solid", start_color=color)
				print(f"-ve sequence image Generated for lane{lane_info}")

				try:
					os.remove(f"{delete_filename}")
					os.remove(f"{delete_filename}.csv")
					os.remove(f"{delete_filename}.log2")
				except:
					print("Some exceptions occurred")

				wb_pos.save(filename = os.path.join(RESULTDIR,args.path,Excel_file_positive))
				wb_neg.save(filename = os.path.join(RESULTDIR,args.path,Excel_file_negative))
		exit(0)
